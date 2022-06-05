
from datetime import datetime
from openpyxl import load_workbook
from sys import argv
from fbGeneratorFillFunctions import fillNetworkFunction, fillTitleFunction, fillStatFunction, fillNetworkFunction2
from excelFunctions import convertExcelColumn

# declare constants 
FirstRow = 2 
unitListColumn = 'A'
tagnameColumn = 'B'
fbTypeColumn = 'C'
parFirstColumn ='D'

# open workbook 
wb = load_workbook(filename=argv[1])   
#sheet = wb.active
sheet = wb["data"]

#grab the unit first cell and put it in the set
unitSet = set()
for row_idx in range(FirstRow, sheet.max_row+1):
    unitListCell = unitListColumn+str(row_idx )
    unitSet.add(sheet[unitListCell].value)
  
#start generator for every Unit   
for unit in unitSet:
    dic = {}

    # create new file
    now = datetime.now()
    newFileName = ".\generatedFiles\\" + str(unit) + now.strftime("_%Y%m%d_%H%M") + ".awl"
    open(newFileName,"w").close()

    #get data to generate STAT 
    for row_idx in range(FirstRow, sheet.max_row+1):
        unitValue = sheet[unitListColumn+str(row_idx )].value
        
        #skip if checked unit does not match 
        if unitValue != unit:
            continue
        
            
        #get parameters
        params=[]
        for column in range (4,sheet.max_column+1):
            parColumn = convertExcelColumn(column)
            params.append(sheet[parColumn+str(row_idx)].value)


        fbType =sheet[fbTypeColumn+str(row_idx)].value
        tagname = sheet[tagnameColumn+str(row_idx)].value  
        
        dic[tagname] = [fbType, params]


    # generate title
    fillTitleFunction(wb, unit,newFileName)

    # generate STAT 

    fillStatFunction (dic, newFileName )

    # generate networks 
    fillNetworkFunction2(wb,dic, "data", unit, newFileName)
    
    #for key in dic.keys():

        #[fb ,par] = dic[key]

        #fillNetworkFunction(wb,fb, "data",unit, key, par,newFileName)    


            






                



            



